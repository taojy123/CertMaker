# -*- coding: utf-8 -*-

from PIL import Image, ImageEnhance, ImageDraw, ImageFont, ImageFilter
from openpyxl import load_workbook
import time
import datetime

def try_str(s):
    try:
        return str(s)
    except:
        return s

def try_date(s):
    try:
        s = s.strftime("%Y年%m月%d日")
        s = s.decode("utf8")
        return s
    except:
        return try_str(s)


def try_print_name(name):
    try:
        print name
    except:
        print ""


print u"========= 欢迎使用证书制作助手 ========="
print u"程序即将启动,请确认:"
print u"1. 证书模板位于 templates 文件夹中"
print u"2. 字体文件位于 fonts 文件夹中"
print u"3. 待添加信息数据位于 data.xlsx 文件中"
print u"4. 存在名为 output 的文件夹(用于存储生成的图片)"
print u"按下回车键开始运行.."
raw_input()



simsun_large = ImageFont.truetype('fonts/simsun.ttc', 72, 0)
simsun = ImageFont.truetype('fonts/simsun.ttc', 56, 0)
dotum = ImageFont.truetype('fonts/gulim.ttc', 60, 3)
simhei = ImageFont.truetype('fonts/simhei.ttf', 56)
arial = ImageFont.truetype('fonts/arial.ttf', 60)
fzdbs = ImageFont.truetype('fonts/fzdbs.ttf', 56)

wb = load_workbook(filename='data.xlsx')
sheet_name = wb.get_sheet_names()[0]
sheet = wb.get_sheet_by_name(sheet_name)

n = 1
while True:
    n += 1
    if not sheet.cell('A' + str(n)).value:
        break
    tp_id = try_str(sheet.cell('A' + str(n)).value)
    name = try_str(sheet.cell('B' + str(n)).value)
    wc_id = try_str(sheet.cell('C' + str(n)).value)
    cert_id = try_str(sheet.cell('D' + str(n)).value)
    product = try_str(sheet.cell('E' + str(n)).value)
    level = try_str(sheet.cell('F' + str(n)).value)
    wc_name = try_str(sheet.cell('G' + str(n)).value)
    validity_from = try_date(sheet.cell('H' + str(n)).value)
    validity_to = try_date(sheet.cell('I' + str(n)).value)

    if tp_id == "1":
        try_print_name(name)
        im = Image.open("templates/1.jpg")
        draw = ImageDraw.Draw(im)
        draw.text((995, 1405), name, font=fzdbs, fill='#000')
        draw.text((1510, 1405), wc_id, font=arial, fill='#000')
        draw.text((1000, 1555), cert_id, font=arial, fill='#000')
        draw.text((1696, 1710), level, font=fzdbs, fill='#000')
        draw.text((710, 1815), u'芙源（亚洲）蜗牛霜%s经销商授权。' % level, font=fzdbs, fill='#000')
        draw.text((860, 2910), u'%s - %s' % (validity_from, validity_to), font=fzdbs, fill='#000')
        im.save("output/%s.jpg" % name)
    elif tp_id == "2":
        try_print_name(name)
        im = Image.open("templates/2.jpg")
        draw = ImageDraw.Draw(im)
        draw.text((989, 1345), name, font=dotum, fill='#000')
        draw.text((1850, 1345), wc_id, font=arial, fill='#000')
        draw.text((938, 1510), cert_id, font=arial, fill='#000')
        draw.text((1657, 1675), level, font=dotum, fill='#000')
        draw.text((620, 1790), u'부행브랜드 %s 중개상권한' % level , font=dotum, fill='#000')
        draw.text((820, 2920), u'%s - %s' % (validity_from, validity_to), font=dotum, fill='#000')
        im.save("output/%s.jpg" % name)
    elif tp_id == "3":
        try_print_name(name)
        im = Image.open("templates/3.jpg")
        draw = ImageDraw.Draw(im)
        draw.text((980, 1480), name, font=fzdbs, fill='#000')
        draw.text((1500, 1472), wc_id, font=arial, fill='#000')
        draw.text((980, 1630), cert_id, font=arial, fill='#000')
        draw.text((1700, 1780), level, font=fzdbs, fill='#000')
        draw.text((700, 1920), u'芙源活泉水系列产品%s经销商授权。' % level, font=fzdbs, fill='#000')
        draw.text((860, 2770), u'%s - %s' % (validity_from, validity_to), font=fzdbs, fill='#000')
        im.save("output/%s.jpg" % name)
    elif tp_id == "4":
        try_print_name(name)
        im = Image.open("templates/4.jpg")
        draw = ImageDraw.Draw(im)
        draw.text((980, 1370), name, font=dotum, fill='#000')
        draw.text((1800, 1370), wc_id, font=arial, fill='#000')
        draw.text((980, 1494), cert_id, font=arial, fill='#000')
        # draw.text((1657, 1675), level, font=dotum, fill='#000')
        draw.text((600, 1880), u'부행브랜드 %s 중개상권한' % level , font=dotum, fill='#000')
        draw.text((800, 2770), u'%s - %s' % (validity_from, validity_to), font=dotum, fill='#000')
        im.save("output/%s.jpg" % name)
    else:
        print u"找不到对应的模板编号:%s" % tp_id

    time.sleep(0.2)

print u"完成!所有生成图片保存在 output 文件夹中"
print u"按下回车键退出..."
raw_input()

#귀사는부원브랜드    상의권한을가진다

